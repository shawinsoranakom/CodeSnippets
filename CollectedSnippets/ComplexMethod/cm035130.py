def write_rows(worksheet, elem, row, column=1):
    """
    Writes every tr child element of elem to a row in the worksheet
    returns the next row after all rows are written
    """
    try_import("openpyxl")
    from openpyxl.cell.cell import MergedCell
    from openpyxl.utils import get_column_letter

    initial_column = column
    for table_row in elem.rows:
        for table_cell in table_row.cells:
            cell = worksheet.cell(row=row, column=column)
            while isinstance(cell, MergedCell):
                column += 1
                cell = worksheet.cell(row=row, column=column)

            colspan = string_to_int(table_cell.element.get("colspan", "1"))
            rowspan = string_to_int(table_cell.element.get("rowspan", "1"))
            if rowspan > 1 or colspan > 1:
                worksheet.merge_cells(
                    start_row=row,
                    start_column=column,
                    end_row=row + rowspan - 1,
                    end_column=column + colspan - 1,
                )

            cell.value = table_cell.value
            table_cell.format(cell)
            min_width = table_cell.get_dimension("min-width")
            max_width = table_cell.get_dimension("max-width")

            if colspan == 1:
                # Initially, when iterating for the first time through the loop, the width of all the cells is None.
                # As we start filling in contents, the initial width of the cell (which can be retrieved by:
                # worksheet.column_dimensions[get_column_letter(column)].width) is equal to the width of the previous
                # cell in the same column (i.e. width of A2 = width of A1)
                width = max(
                    worksheet.column_dimensions[get_column_letter(column)].width or 0,
                    len(table_cell.value) + 2,
                )
                if max_width and width > max_width:
                    width = max_width
                elif min_width and width < min_width:
                    width = min_width
                worksheet.column_dimensions[get_column_letter(column)].width = width
            column += colspan
        row += 1
        column = initial_column
    return row