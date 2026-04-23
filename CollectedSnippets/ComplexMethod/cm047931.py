def _read_xlsx(self, options):
        try:
            from xlrd import xlsx  # noqa: F401, PLC0415
            if xlsx:
                return self._read_xls(options)
        except ImportError:
            pass

        import openpyxl  # noqa: PLC0415
        import openpyxl.cell.cell as types  # noqa: PLC0415
        import openpyxl.styles.numbers as styles  # noqa: PLC0415
        book = openpyxl.load_workbook(io.BytesIO(self.file or b''), data_only=True)
        sheets = options['sheets'] = book.sheetnames
        sheet_name = options['sheet'] = options.get('sheet') or sheets[0]
        sheet = book[sheet_name]
        rows = []
        for rowx, row in enumerate(sheet.rows, 1):
            values = []
            for colx, cell in enumerate(row, 1):
                if cell.data_type == types.TYPE_ERROR:
                    raise ValueError(
                        _("Invalid cell value at row %(row)s, column %(col)s: %(cell_value)s", row=rowx, col=colx, cell_value=cell.value)
                    )

                if cell.value is None:
                    values.append('')
                elif isinstance(cell.value, float):
                    if cell.value % 1 == 0:
                        values.append(str(int(cell.value)))
                    else:
                        values.append(str(cell.value))
                elif cell.is_date:
                    d_fmt = styles.is_datetime(cell.number_format)
                    if d_fmt == "datetime":
                        values.append(cell.value)
                    elif d_fmt == "date":
                        values.append(cell.value.date())
                    else:
                        raise ValueError(
                        _("Invalid cell format at row %(row)s, column %(col)s: %(cell_value)s, with format: %(cell_format)s, as (%(format_type)s) formats are not supported.", row=rowx, col=colx, cell_value=cell.value, cell_format=cell.number_format, format_type=d_fmt)
                        )
                else:
                    values.append(str(cell.value))

            if any(x and (not isinstance(x, str) or x.strip()) for x in values):
                rows.append(values)
        return sheet.max_row, rows