def convert_file(self, file_path: Path, **kwargs) -> ConvertResult:
        try:
            import openpyxl
            from openpyxl.drawing.spreadsheet_drawing import (
                OneCellAnchor,
                TwoCellAnchor,
            )
        except ImportError:
            raise RuntimeError(
                "XLSX conversion requires openpyxl: pip install paddleocr[doc2md]"
            )

        sheet_name: Optional[str] = kwargs.get("sheet_name", None)
        max_rows: Optional[int] = kwargs.get("max_rows", None)
        extract_drawings: bool = kwargs.get("extract_drawings", True)

        import zipfile

        # read_only=False is required to access merged_cells
        wb = openpyxl.load_workbook(str(file_path), read_only=False, data_only=True)
        sheets_md = []
        images: dict = {}
        image_counter = 0

        target_sheets = [sheet_name] if sheet_name else wb.sheetnames

        with zipfile.ZipFile(str(file_path), "r") as _zf:
            for sname in target_sheets:
                if sname not in wb.sheetnames:
                    continue
                ws = wb[sname]

                # Total sheet width in EMU, used for image percentage calculation
                sheet_width_emu = _get_sheet_width_emu(ws, openpyxl)

                # Floating image map: (0-based row, 0-based col) -> [Image, ...]
                image_map: dict = {}
                for img in getattr(ws, "_images", []):
                    anchor = img.anchor
                    if isinstance(anchor, (OneCellAnchor, TwoCellAnchor)):
                        r, c = anchor._from.row, anchor._from.col
                        image_map.setdefault((r, c), []).append(img)
                    elif isinstance(anchor, str):
                        try:
                            from openpyxl.utils import coordinate_to_tuple

                            r, c = coordinate_to_tuple(anchor)
                            image_map.setdefault((r - 1, c - 1), []).append(img)
                        except Exception:
                            pass

                # Merged cell map: (row, col) -> MergedCellRange
                merge_map = {}
                for mr in ws.merged_cells.ranges:
                    for r in range(mr.min_row, mr.max_row + 1):
                        for c in range(mr.min_col, mr.max_col + 1):
                            merge_map[(r, c)] = mr

                # Trim surrounding empty rows/columns
                bounds = _find_data_bounds(ws, image_map, max_rows)
                if bounds is None:
                    continue
                data_min_row, data_max_row, data_min_col, data_max_col = bounds

                html_parts = ["<table>"]
                for row_idx in range(data_min_row, data_max_row + 1):
                    html_parts.append("<tr>")
                    for col_idx in range(data_min_col, data_max_col + 1):
                        cell = ws.cell(row_idx, col_idx)
                        mr = merge_map.get((row_idx, col_idx))
                        # Skip non-origin cells in a merged range
                        if mr and (row_idx, col_idx) != (mr.min_row, mr.min_col):
                            continue
                        tag = "th" if row_idx == data_min_row else "td"
                        attrs = ""
                        if mr:
                            cs = (
                                min(mr.max_col, data_max_col)
                                - max(mr.min_col, data_min_col)
                                + 1
                            )
                            rs = (
                                min(mr.max_row, data_max_row)
                                - max(mr.min_row, data_min_row)
                                + 1
                            )
                            if cs > 1:
                                attrs += f' colspan="{cs}"'
                            if rs > 1:
                                attrs += f' rowspan="{rs}"'

                        # Cell text
                        value = cell.value
                        text = str(value) if value is not None else ""
                        # Cell-level font formatting (bold/italic/underline/strikethrough)
                        if text:
                            try:
                                font = cell.font
                                if font.bold:
                                    text = f"<b>{text}</b>"
                                if font.italic:
                                    text = f"<i>{text}</i>"
                                if font.underline:
                                    text = f"<u>{text}</u>"
                                if font.strike:
                                    text = f"<del>{text}</del>"
                                vert_align = font.vertAlign
                                if vert_align == "superscript":
                                    text = f"<sup>{text}</sup>"
                                elif vert_align == "subscript":
                                    text = f"<sub>{text}</sub>"
                            except Exception:
                                pass
                        # Hyperlink wrapping
                        if text:
                            try:
                                hl = cell.hyperlink
                                if hl and hl.target:
                                    text = f'<a href="{hl.target}">{text}</a>'
                            except Exception:
                                pass

                        # Floating images
                        cell_images = image_map.get((row_idx - 1, col_idx - 1), [])
                        img_html = ""
                        for img_obj in cell_images:
                            image_counter += 1
                            ext = (img_obj.format or "png").lower()
                            filename = f"image{image_counter}.{ext}"
                            rel_path = f"images/{filename}"
                            try:
                                ref = img_obj.ref
                                if isinstance(ref, BytesIO):
                                    ref.seek(0)
                                    images[rel_path] = ref.read()
                                else:
                                    images[rel_path] = img_obj._data()
                                cx_emu = _get_image_cx(img_obj.anchor)
                                if cx_emu and sheet_width_emu:
                                    pct = min(
                                        round(cx_emu / sheet_width_emu * 100), 100
                                    )
                                    img_html += (
                                        f'<img src="images/{filename}" width="{pct}%">'
                                    )
                                else:
                                    img_html += f'<img src="images/{filename}">'
                            except Exception:
                                pass

                        cell_content = img_html + text if img_html else text
                        html_parts.append(f"<{tag}{attrs}>{cell_content}</{tag}>")
                    html_parts.append("</tr>")
                html_parts.append("</table>")

                table_html = "\n".join(html_parts)

                sheet_parts = [f"## {sname}\n\n{table_html}"]
                if extract_drawings:
                    sheet_idx = list(wb.sheetnames).index(sname)
                    for latex in _extract_drawing_math(_zf, sheet_idx):
                        sheet_parts.append(f"\n$$\n{latex}\n$$")
                sheets_md.append("\n".join(sheet_parts))

        sheet_names = list(wb.sheetnames)
        wb.close()

        md_text = "\n\n".join(sheets_md)

        return ConvertResult(
            markdown=md_text,
            images=images,
            metadata={
                "format": "XLSX",
                "sheet_count": len(sheet_names),
                "sheet_names": sheet_names,
            },
        )