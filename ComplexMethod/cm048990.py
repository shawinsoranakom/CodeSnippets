def _index_xlsx(self, bin_data):
        '''Index Microsoft .xlsx documents'''

        try:
            from openpyxl import load_workbook  # noqa: PLC0415
            logging.getLogger("openpyxl").setLevel(logging.CRITICAL)
        except ImportError:
            _logger.info('openpyxl is not installed.')
            return ""

        f = io.BytesIO(bin_data)
        all_sheets = []
        try:
            with warnings.catch_warnings():
                warnings.simplefilter("ignore")
                workbook = load_workbook(f, data_only=True, read_only=True)
                for sheet in workbook.worksheets:
                    sheet_name = sheet.title
                    sheet_name_escaped = _csv_escape(sheet_name)
                    sheet_rows = []
                    for row in sheet.iter_rows(values_only=True):
                        if not any(row):
                            continue
                        row_cells = [sheet_name_escaped] + [
                            _csv_escape(str(cell) if cell is not None else '') for cell in row
                        ]
                        sheet_rows.append(','.join(row_cells))
                    sheet_data = '\n'.join(sheet_rows)
                    if sheet_data:
                        all_sheets.append(sheet_data)
        except Exception:  # noqa: BLE001
            pass

        all_sheets_str = '\n\n'.join(all_sheets)
        return _clean_text_content(all_sheets_str)